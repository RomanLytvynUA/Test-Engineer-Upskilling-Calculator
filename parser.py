import json
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import re
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

tables = ['Base Roles', 'Competency template', 'Tools\\Frameworks']
# tables = ['Competency template']
file = openpyxl.load_workbook("file.xlsx")
sh = file.active
trainings_sh = file["Trainings"]


def generate_trainings_json():
    data = []

    for row in range(3, trainings_sh.max_row + 1):

        # Stop if table ended (col value is none)
        if trainings_sh.cell(row=row, column=2).value is None: break

        data.append({'name': trainings_sh.cell(row=row, column=2).value,
                     'link': trainings_sh.cell(row=row, column=3).value,
                     'level': trainings_sh.cell(row=row, column=4).value,
                     'duration': trainings_sh.cell(row=row, column=5).value,
                     })
    return data


def generate_json(table_name):
    data = {}

    table = find_cel(table_name)

    if table == (0, 0): return {}

    table_entities_count = 1
    while True:
        entity = sh.cell(row=table[0] + table_entities_count, column=table[1]).value

        # Stop if table ended (cel is name of a table OR cel in empty, and it's last table)
        if entity in tables or (entity is None and tables[-1] == table_name): break

        # Only add a fields if cel is not empty
        if entity is not None: data[entity] = get_fields(table[0] + table_entities_count)
        table_entities_count += 1
    return data


def find_cel(content):
    for row in range(1, sh.max_row + 1):
        for col in range(1, sh.max_column + 1):
            value = sh.cell(row=row, column=col).value
            if value is not None and value.lower() == content.lower():
                return row, col
    return 0, 0


def get_fields(row, col_start=3):
    fields = []
    for col in range(col_start, sh.max_column + 1):
        field_value = sh.cell(row=row, column=col).value

        # Stop if row ended (cel is empty)
        if field_value is None: break

        if re.match(r'=CONCATENATE\("Select ", (\w+)\)', field_value):
            field_data = {}
            coordinates = coordinate_from_string(re.match(r'=CONCATENATE\("Select ", (\w+)\)', field_value).group(1))

            field_data['name'] = sh.cell(row=coordinates[1], column=column_index_from_string(coordinates[0])).value
            field_data['options'] = get_fields(coordinates[1])

            fields.append(field_data)
        elif field_value[:6] == "Select":
            field_data = {'name': field_value[7:]}

            field_options = find_cel(field_data['name'])[0]

            # Check if field options are defined in the file
            if field_options != 0:
                field_data['options'] = get_fields(field_options)
                fields.append(field_data)
            else:
                fields.append(field_data['name'])
        else:
            fields.append(field_value)
    return fields

generate_trainings_json()
for table in tables:
    file_name = table

    restricted_characters = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    for letter in file_name:
        if letter in restricted_characters: file_name = table.replace(letter, '_')

    with open(f'./data/{file_name}.json', 'w') as f:
        json.dump(generate_json(table), f)
    
with open(f'./data/trainings.json', 'w') as f:
    json.dump(generate_trainings_json(), f)
# type: ignore
